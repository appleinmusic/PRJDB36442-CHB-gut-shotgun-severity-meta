#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Extract per-sample group labels from NGDC GSA export Excel (CRA023641.xlsx)."
    )
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=Path("data/metadata/ngdc/CRA023641.xlsx"),
        help="Path to CRA023641.xlsx (default: data/metadata/ngdc/CRA023641.xlsx).",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=Path("results/feasibility/PRJCA037061_sample_groups.tsv"),
        help="Output TSV path (default: results/feasibility/PRJCA037061_sample_groups.tsv).",
    )
    args = ap.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    if "Sample" not in wb.sheetnames:
        raise SystemExit(f"Missing sheet 'Sample' in {args.xlsx}")

    ws = wb["Sample"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"No rows in Sample sheet: {args.xlsx}")

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    idx = {name: i for i, name in enumerate(header) if name}
    required = ["Sample name", "Accession", "Public description", "Project accession"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise SystemExit(f"Missing columns in Sample sheet: {missing}")

    out_rows: list[dict[str, str]] = []
    for r in rows[1:]:
        if not r or all(v is None for v in r):
            continue
        bf = (r[idx["Sample name"]] or "").strip()
        samc = (r[idx["Accession"]] or "").strip()
        desc = (r[idx["Public description"]] or "").strip()
        prj = (r[idx["Project accession"]] or "").strip()
        group = ""
        if desc.lower().startswith("group "):
            group = desc.split(" ", 1)[1].strip()
        out_rows.append(
            {
                "project_accession": prj,
                "sample_name": bf,
                "ngdc_sample_accession": samc,
                "public_description": desc,
                "group": group,
                "source_xlsx": str(args.xlsx),
            }
        )

    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "project_accession",
                "sample_name",
                "ngdc_sample_accession",
                "public_description",
                "group",
                "source_xlsx",
            ],
            delimiter="\t",
        )
        w.writeheader()
        w.writerows(out_rows)

    print(str(args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

